# Copyright (c) 2023, frappe and contributors
# For license information, please see license.txt

# import frappe
from frappe.model.document import Document
import os
import openpyxl
import frappe

class StockBalanceForm(Document):
	pass




def is_excel_file(file_name):
    _, ext = os.path.splitext(file_name)
    return ext == '.xlsx'

def create_or_update_item(item_code, item_name, uom, warehouse_name, balance_qty, machine_type):
    item_list = frappe.db.get_list('Item', pluck='name')
    uom_list_lower = [uom.lower() for uom in frappe.db.get_list('UOM', pluck='name')]

    if item_code not in item_list and not frappe.db.exists("Item", item_code):
        item_doc = frappe.new_doc("Item")
        item_doc.item_code = item_code
        item_doc.item_name = item_name
        item_doc.machine_type = machine_type
        item_doc.item_group = "Products"
        item_doc.has_batch_no = 0

        if uom.lower() not in uom_list_lower:
            uom_doc = frappe.new_doc("UOM")
            uom_doc.uom_name = uom
            uom_doc.save()
            item_doc.stock_uom = uom
        else:
            item_doc.stock_uom = uom

        item_doc.save()

    update_stock_balance(item_code, warehouse_name, balance_qty)

def update_stock_balance(item_code, warehouse_name, balance_qty):
    sbf = frappe.db.get_list('SBF TEST', {'item_code': item_code, 'warehouse': warehouse_name})
    if not sbf:
        # Create a new SBF TEST entry
        stc_bal = frappe.new_doc("SBF TEST")
        stc_bal.item_code = item_code
        stc_bal.warehouse = warehouse_name
    else:
        # Update the existing SBF TEST entry
        stc_bal = frappe.get_doc("SBF TEST", sbf[0].name)

    stc_bal.balance_qty = balance_qty
    stc_bal.save()

@frappe.whitelist()
def make_entries(file_name, warehouse_name, doc):
    try:
        if not is_excel_file(file_name):
            frappe.throw('Please upload .xlsx format')

        media_file_path = frappe.get_doc("File", {"file_url": file_name}).get_full_path()
        ise_sheet = openpyxl.load_workbook(media_file_path)
        ise_file = ise_sheet.active

        # Create a list to track existing item codes from the Excel data
        existing_item_codes = []
        l=[]

        for row in range(2, ise_file.max_row + 1):
            item_code = ise_file.cell(row=row, column=2).value
            item_name = ise_file.cell(row=row, column=3).value
            uom = ise_file.cell(row=row, column=4).value
            balance_qty = ise_file.cell(row=row, column=5).value
            machine_type = ise_file.cell(row=row, column=6).value

            create_or_update_item(item_code, item_name, uom, warehouse_name, balance_qty, machine_type)

            existing_item_codes.append(item_code)

        # Set balance_qty to 0 for items not found in the uploaded Excel data
        frappe.log_error(f'items,{existing_item_codes}')
        sbf_entries = frappe.db.get_list('SBF TEST', {'warehouse': warehouse_name}, 'item_code', pluck='item_code')
        for k in sbf_entries:
            # item_code = entry.get('item_code')
            frappe.log_error(f'item2,{k}')
            if k not in existing_item_codes:
                l.append(k)
        frappe.log_error(f"l,{l}")
        

                

        frappe.msgprint("SBF TEST is Successfully Created")
    except Exception as e:
        frappe.log_error(f"Error in make_entries: {str(e)}")
        frappe.throw("An error occurred while processing the file.")
